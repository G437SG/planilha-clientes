"""
Formulário de Projeto Arquitetônico - Aplicativo para coleta e exportação
de informações de clientes e projetos arquitetônicos.
"""
import os
import sys
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox, filedialog
from typing import Dict, List, Tuple, Optional, Any, Union

from fpdf import FPDF
from openpyxl import Workbook
from PIL import ImageGrab

from ui import UI


class AppLogic:
    """Lógica principal do aplicativo de formulário de projetos arquitetônicos."""

    def __init__(self) -> None:
        """Inicializa a classe AppLogic."""
        self.fields: Dict[str, ttk.Entry] = {}
        self.checks: Dict[str, tk.BooleanVar] = {}
        self.radio_vars: Dict[str, Union[tk.StringVar, tk.BooleanVar]] = {}
        self.demandas_entries: List[Tuple[ttk.Entry, ttk.Entry]] = []
        self.detalhamento_vars: Dict[str, tk.BooleanVar] = {}
        self.export_path = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop"))
        self.root: Optional[tk.Tk] = None
        self.other_imovel_frame: Optional[ttk.Frame] = None
        self.construction_frame: Optional[ttk.Frame] = None
        self.detalhamento_frame: Optional[ttk.Frame] = None
        self.demandas_frame: Optional[ttk.Frame] = None

    def set_root(self, root: tk.Tk) -> None:
        """Define a janela raiz do aplicativo.
        
        Args:
            root: A janela raiz Tkinter.
        """
        self.root = root

    def add_client_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de informações do cliente ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=10)
        ttk.Label(section_frame, text="INFORMAÇÕES DO CLIENTE", style="Header.TLabel").pack(fill=tk.X)

        # Aumenta a largura das caixas de texto e ajusta os rótulos
        self.add_labeled_entry(section_frame, "Nome completo:", "nome", validate="nome", width=50)
        self.add_labeled_entry(section_frame, "Telefone:", "telefone", validate="telefone", width=50)
        self.add_labeled_entry(section_frame, "Email:", "email", width=50)
        self.add_labeled_entry(section_frame, "CNPJ/CPF:", "cnpj", width=50)
        self.add_labeled_entry(section_frame, "Responsável legal:", "responsavel", width=50)
        self.add_labeled_entry(section_frame, "Telefone do Responsável:", "telefone_responsavel", validate="telefone", width=50)
        self.add_labeled_entry(section_frame, "Endereço:", "endereco", width=50)
        self.add_labeled_entry(section_frame, "CEP:", "cep", validate="cep", width=50)

    def add_property_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de informações do imóvel ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=10)
        ttk.Label(section_frame, text="INFORMAÇÕES DO IMÓVEL", style="Header.TLabel").pack(fill=tk.X)

        type_frame = ttk.LabelFrame(section_frame, text="Tipo de Imóvel")
        type_frame.pack(fill=tk.X, pady=5, padx=5)

        self.radio_vars["tipo_imovel"] = tk.StringVar(value="comercial")
        types = [
            ("Sala Comercial", "comercial"),
            ("Casa", "casa"),
            ("Prédio", "predio"),
            ("Outro", "outro")
        ]
        for text, value in types:
            rb = ttk.Radiobutton(
                type_frame, 
                text=text, 
                value=value, 
                variable=self.radio_vars["tipo_imovel"], 
                command=self.update_other_imovel_input
            )
            rb.pack(side=tk.LEFT, padx=10)

        # Campo de entrada para "Outro"
        self.other_imovel_frame = ttk.Frame(section_frame)
        self.other_imovel_frame.pack(fill=tk.X, pady=5, padx=5)
        ttk.Label(
            self.other_imovel_frame, 
            text="Especifique o tipo de imóvel:", 
            width=25, 
            anchor="w"
        ).pack(side=tk.LEFT)
        self.other_imovel_entry = ttk.Entry(self.other_imovel_frame, width=50)
        self.other_imovel_entry.pack(side=tk.LEFT, padx=5)

        # Inicialmente, oculta o campo de entrada
        self.other_imovel_frame.pack_forget()

        # Campo de Metragem Quadrada
        metragem_frame = ttk.Frame(section_frame)
        metragem_frame.pack(fill=tk.X, pady=5, padx=5)
        ttk.Label(metragem_frame, text="Metragem Quadrada:", width=25, anchor="w").pack(side=tk.LEFT)
        self.fields["metragem"] = ttk.Entry(metragem_frame, width=50)
        self.fields["metragem"].pack(side=tk.LEFT, padx=5)

        # Campo de tipo de construção
        self.construction_frame = ttk.Frame(section_frame)
        self.construction_frame.pack(fill=tk.X, pady=5)

        self.radio_vars["tipo_construcao"] = tk.StringVar(value="reforma")
        ttk.Radiobutton(
            self.construction_frame, 
            text="Reforma", 
            value="reforma", 
            variable=self.radio_vars["tipo_construcao"]
        ).pack(side=tk.LEFT, padx=10)
        ttk.Radiobutton(
            self.construction_frame, 
            text="Construção do Zero", 
            value="nova", 
            variable=self.radio_vars["tipo_construcao"]
        ).pack(side=tk.LEFT, padx=10)

        self.update_construction_options()
        self.radio_vars["tipo_imovel"].trace_add("write", lambda *args: self.update_construction_options())

    def update_other_imovel_input(self) -> None:
        """Atualiza a visibilidade do campo de entrada 'Outro' para tipo de imóvel."""
        tipo = self.radio_vars["tipo_imovel"].get()
        if tipo == "outro":
            self.other_imovel_frame.pack(fill=tk.X, pady=5, padx=5)
        else:
            self.other_imovel_frame.pack_forget()

    def update_construction_options(self) -> None:
        """Atualiza a visibilidade das opções de construção."""
        tipo = self.radio_vars["tipo_imovel"].get()
        if tipo in ["casa", "predio"]:
            self.construction_frame.pack(fill=tk.X, pady=5)
        else:
            self.construction_frame.pack_forget()

    def add_scope_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de escopo do projeto ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=10)
        ttk.Label(section_frame, text="ESCOPO", style="Header.TLabel").pack(fill=tk.X)

        # Projeto de Arquitetura
        ttk.Label(section_frame, text="Projeto Arquitetura:", style="Section.TLabel").pack(fill=tk.X, pady=5)
        architecture_options = [
            ("Layout", "layout"),
            ("3D", "3d"),
            ("Detalhamento", "detalhamento")
        ]

        for text, key in architecture_options:
            if key == "detalhamento":
                # Cria um contêiner para "Detalhamento" e suas subopções
                detalhamento_container = ttk.Frame(section_frame)
                detalhamento_container.pack(fill=tk.X, padx=20, pady=5)

                var = tk.BooleanVar(value=True)
                self.checks[key] = var
                cb = ttk.Checkbutton(
                    detalhamento_container, 
                    text=text, 
                    variable=var, 
                    command=self.update_detalhamento_options
                )
                cb.pack(anchor=tk.W)

                # Subopções de Detalhamento (presas logo abaixo de "Detalhamento")
                self.detalhamento_frame = ttk.Frame(detalhamento_container)
                self.detalhamento_frame.pack(fill=tk.X, padx=20, pady=5)

                detalhamento_options = [
                    "Marcenaria",
                    "Detalhamento Áreas Molhadas",
                    "Forro",
                    "Iluminação",
                    "Tomadas",
                    "Pisos",
                    "Executiva",
                    "Layout",
                    "Demolir e Construir",
                    "Apresentação"
                ]
                self.detalhamento_vars = {}
                for option in detalhamento_options:
                    var = tk.BooleanVar(value=True)  # Selecionado como padrão
                    self.detalhamento_vars[option] = var
                    ttk.Checkbutton(self.detalhamento_frame, text=option, variable=var).pack(anchor=tk.W)

                # Inicialmente, o frame de detalhamento estará visível (porque value=True)
                self.update_detalhamento_options()
            else:
                var = tk.BooleanVar(value=True)
                self.checks[key] = var
                cb = ttk.Checkbutton(section_frame, text=text, variable=var)
                cb.pack(anchor=tk.W, padx=20)

        # Projetos Complementares
        ttk.Label(section_frame, text="Projetos Complementares:", style="Section.TLabel").pack(fill=tk.X, pady=5)
        complementary_options = [
            ("Ar Condicionado", "ar_condicionado"),
            ("Elétrica", "eletrica"),
            ("Dados e Voz", "dados_voz"),
            ("Hidráulica", "hidraulica"),
            ("CFTV", "cftv"),
            ("Alarme", "alarme"),
            ("Incêndio", "incendio")
        ]
        for text, key in complementary_options:
            var = tk.BooleanVar(value=True)
            self.checks[key] = var
            ttk.Checkbutton(section_frame, text=text, variable=var).pack(anchor=tk.W, padx=20)

    def update_detalhamento_options(self) -> None:
        """Atualiza a visibilidade das opções de detalhamento."""
        if self.checks.get("detalhamento", tk.BooleanVar()).get():
            self.detalhamento_frame.pack(fill=tk.X, padx=20, pady=5)
        else:
            self.detalhamento_frame.pack_forget()

    def add_demands_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de demandas do projeto ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=10)

        ttk.Label(section_frame, text="DEMANDAS DO PROJETO", style="Header.TLabel").pack(fill=tk.X)
        header_frame = ttk.Frame(section_frame)
        header_frame.pack(fill=tk.X, pady=5)
        ttk.Label(header_frame, text="Nome", width=30, anchor="w").pack(side=tk.LEFT, padx=5)
        ttk.Label(header_frame, text="Descrição", width=50, anchor="w").pack(side=tk.LEFT, padx=5)
        ttk.Label(header_frame, text="Ações", width=10, anchor="w").pack(side=tk.LEFT, padx=5)

        self.demandas_frame = ttk.Frame(section_frame)
        self.demandas_frame.pack(fill=tk.X, pady=5)
        self.add_demanda_row()

    def add_demanda_row(self) -> None:
        """Adiciona uma nova linha de demanda ao formulário."""
        row_frame = ttk.Frame(self.demandas_frame)
        row_frame.pack(fill=tk.X, pady=2)

        nome_entry = ttk.Entry(row_frame, width=30)
        nome_entry.pack(side=tk.LEFT, padx=5)
        nome_entry.bind("<FocusOut>", lambda e: self.check_and_add_row())

        descricao_entry = ttk.Entry(row_frame, width=50)
        descricao_entry.pack(side=tk.LEFT, padx=5)
        descricao_entry.bind("<FocusOut>", lambda e: self.check_and_add_row())

        self.demandas_entries.append((nome_entry, descricao_entry))

        action_frame = ttk.Frame(row_frame)
        action_frame.pack(side=tk.LEFT, padx=5)

        ttk.Button(
            action_frame, 
            text="🗑️", 
            width=3, 
            command=lambda: self.remove_demanda_row(row_frame, (nome_entry, descricao_entry))
        ).pack(side=tk.LEFT, padx=2)
        ttk.Button(action_frame, text="➕", width=3, command=self.add_demanda_row).pack(side=tk.LEFT, padx=2)

    def remove_demanda_row(self, row_frame: ttk.Frame, entries: Tuple[ttk.Entry, ttk.Entry]) -> None:
        """Remove uma linha de demanda do formulário.
        
        Args:
            row_frame: O frame da linha a ser removida.
            entries: Tupla contendo as entradas de nome e descrição da linha.
        """
        if len(self.demandas_entries) > 1:  # Garantir que pelo menos uma linha permaneça
            self.demandas_entries.remove(entries)
            row_frame.destroy()

    def check_and_add_row(self) -> None:
        """Verifica se a última linha de demanda foi preenchida e adiciona uma nova se necessário."""
        if not self.demandas_entries:
            return
        last_row = self.demandas_entries[-1]
        if all(entry.get().strip() for entry in last_row):
            self.add_demanda_row()

    def add_deadlines_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de prazos do projeto ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=10)
        ttk.Label(section_frame, text="PRAZOS DO PROJETO", style="Header.TLabel").pack(fill=tk.X)

        deadlines = [
            ("Levantamento:", "levantamento"),
            ("Layout:", "layout"),
            ("Modelagem 3D:", "modelagem_3d"),
            ("Projeto Executivo:", "projeto_executivo"),
            ("Complementares:", "complementares")
        ]
        for label, key in deadlines:
            self.add_labeled_entry(section_frame, label, key, validate="dias", width=10, suffix="DIAS")

    def add_buttons_section(self, frame: ttk.Frame) -> None:
        """Adiciona a seção de botões ao formulário.
        
        Args:
            frame: O frame onde a seção será adicionada.
        """
        section_frame = ttk.Frame(frame)
        section_frame.pack(fill=tk.X, pady=20)

        # Campo de seleção de pasta
        ttk.Label(section_frame, text="Pasta de Exportação:").pack(side=tk.LEFT, padx=5)
        ttk.Entry(section_frame, textvariable=self.export_path, width=50).pack(side=tk.LEFT, padx=5)
        ttk.Button(
            section_frame, 
            text="Selecionar Pasta", 
            command=self.select_export_path
        ).pack(side=tk.LEFT, padx=5)

        # Botões abaixo do campo de seleção de pasta
        btn_frame = ttk.Frame(frame)  # Novo frame para os botões
        btn_frame.pack(fill=tk.X, pady=10)  # Posicionado abaixo do campo de seleção de pasta
        ttk.Button(btn_frame, text="Exportar para PDF", command=self.export_to_pdf).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Exportar para Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=10)

    def select_export_path(self) -> None:
        """Abre um diálogo para selecionar a pasta de exportação."""
        folder = filedialog.askdirectory()
        if folder:
            self.export_path.set(folder)

    def add_labeled_entry(self, parent: ttk.Frame, label: str, key: str, validate: Optional[str] = None, 
                         date: bool = False, width: int = 40, suffix: Optional[str] = None) -> None:
        """Adiciona um campo de entrada rotulado ao formulário.
        
        Args:
            parent: O frame pai onde o campo será adicionado.
            label: O texto do rótulo.
            key: A chave para armazenar o campo no dicionário self.fields.
            validate: O tipo de validação a ser aplicada ao campo.
            date: Se o campo é uma data.
            width: A largura do campo de entrada.
            suffix: Texto a ser exibido após o campo.
        """
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=2)
        ttk.Label(frame, text=label, width=25, anchor="w").pack(side=tk.LEFT)
        entry = ttk.Entry(frame, width=width)
        entry.pack(side=tk.LEFT, padx=5)

        if suffix:
            ttk.Label(frame, text=suffix).pack(side=tk.LEFT, padx=5)

        if validate == "dias":
            # Validação para garantir que apenas números sejam inseridos
            def validate_days(event):
                value = entry.get().strip()
                if value and not value.isdigit():
                    messagebox.showerror("Erro", f"Valor inválido: {value}. Insira apenas números.")
                    entry.delete(0, tk.END)

            entry.bind("<FocusOut>", validate_days)

        self.fields[key] = entry

    def save_data(self) -> None:
        """Salva os dados do formulário."""
        data = {key: field.get() for key, field in self.fields.items()}
        messagebox.showinfo("Salvar Dados", f"Dados salvos com sucesso:\n{data}")

    def clear_form(self) -> None:
        """Limpa todos os campos do formulário."""
        if not messagebox.askyesno("Confirmar", "Deseja realmente limpar todos os campos?"):
            return
        for field in self.fields.values():
            if isinstance(field, tk.Entry):
                field.delete(0, tk.END)
        for var in self.checks.values():
            var.set(False)
        for var in self.radio_vars.values():
            if isinstance(var, tk.StringVar):
                var.set("")
            elif isinstance(var, tk.BooleanVar):
                var.set(False)
        for nome_entry, descricao_entry in self.demandas_entries:
            nome_entry.delete(0, tk.END)
            descricao_entry.delete(0, tk.END)
        messagebox.showinfo("Limpar Formulário", "Todos os campos foram limpos.")

    def _get_export_filename(self, extension: str) -> Tuple[str, str]:
        """Gera um nome de arquivo para exportação.
        
        Args:
            extension: A extensão do arquivo.
            
        Returns:
            Tupla contendo o nome do arquivo e o caminho completo.
        """
        nome_cliente = self.fields.get("nome", ttk.Entry()).get().strip() or "Cliente"
        data_atual = datetime.now().strftime("%d-%m-%Y")
        filename = f"{nome_cliente} - RELATÓRIO - {data_atual}.{extension}"
        file_path = os.path.join(self.export_path.get(), filename)
        return filename, file_path

    def export_to_pdf(self) -> None:
        """Exporta os dados do formulário para um arquivo PDF."""
        if not self.root or not self.export_path.get():
            messagebox.showerror("Erro", "Selecione uma pasta de exportação antes de continuar.")
            return

        # Nome do arquivo e caminho
        _, pdf_path = self._get_export_filename("pdf")

        # Configuração do PDF
        pdf = FPDF()
        pdf.add_page()

        # Cabeçalho com a logo
        pdf.set_font("Arial", size=10)
        base_path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
        logo_path = os.path.join(base_path, 'logo_empresa.png')
        pdf.image(logo_path, x=10, y=8, w=30)
        pdf.cell(0, 8, "Relatório do Formulário", ln=True, align="C")
        pdf.ln(5)

        # INFORMAÇÕES DO CLIENTE
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "INFORMAÇÕES DO CLIENTE:", ln=True)
        pdf.set_font("Arial", size=9)
        client_info = {
            "Nome completo": self.fields.get("nome", ttk.Entry()).get(),
            "Telefone": self.fields.get("telefone", ttk.Entry()).get(),
            "Email": self.fields.get("email", ttk.Entry()).get(),
            "CNPJ/CPF": self.fields.get("cnpj", ttk.Entry()).get(),
            "Responsável legal": self.fields.get("responsavel", ttk.Entry()).get(),
            "Telefone do Responsável": self.fields.get("telefone_responsavel", ttk.Entry()).get(),
            "Endereço": self.fields.get("endereco", ttk.Entry()).get(),
            "CEP": self.fields.get("cep", ttk.Entry()).get()
        }
        for label, value in client_info.items():
            pdf.cell(0, 6, f"{label}: {value}", ln=True)

        pdf.ln(5)

        # INFORMAÇÕES DO IMÓVEL
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "INFORMAÇÕES DO IMÓVEL:", ln=True)
        pdf.set_font("Arial", size=9)
        property_info = {
            "Tipo de Imóvel": self.radio_vars.get("tipo_imovel", tk.StringVar()).get(),
            "Tipo de Construção": self.radio_vars.get("tipo_construcao", tk.StringVar()).get(),
            "Metragem Quadrada": f"{self.fields.get('metragem', ttk.Entry()).get()} m²"
        }
        for label, value in property_info.items():
            pdf.cell(0, 6, f"{label}: {value}", ln=True)

        pdf.ln(5)

        # PRAZOS DO PROJETO
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "PRAZOS DO PROJETO:", ln=True)
        pdf.set_font("Arial", size=9)
        deadlines = {
            "Levantamento": "levantamento",
            "Layout": "layout",
            "Modelagem 3D": "modelagem_3d",
            "Projeto Executivo": "projeto_executivo",
            "Complementares": "complementares"
        }
        for label, key in deadlines.items():
            days_value = self.fields.get(key, ttk.Entry()).get()
            pdf.cell(0, 6, f"{label}: {days_value} DIAS" if days_value else f"{label}: Não informado", ln=True)

        pdf.ln(5)

        # DEMANDAS DO PROJETO
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "DEMANDAS DO PROJETO:", ln=True)
        pdf.set_font("Arial", size=9)
        for nome_entry, descricao_entry in self.demandas_entries:
            nome = nome_entry.get().strip()
            descricao = descricao_entry.get().strip()
            if nome or descricao:
                pdf.cell(0, 6, f"Nome: {nome} - Descrição: {descricao}", ln=True)

        pdf.ln(5)

        # PROJETO DE ARQUITETURA
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "PROJETO DE ARQUITETURA:", ln=True)
        pdf.set_font("Arial", size=9)
        architecture_options = [
            ("Layout", "layout"),
            ("3D", "3d"),
            ("Detalhamento", "detalhamento")
        ]
        for text, key in architecture_options:
            if self.checks.get(key, tk.BooleanVar()).get():
                pdf.cell(0, 5, f"- {text}", ln=True)

        # Subopções de Detalhamento
        if self.checks.get("detalhamento", tk.BooleanVar()).get():
            detalhamento_options = [
                "Marcenaria",
                "Detalhamento Áreas Molhadas",
                "Forro",
                "Iluminação",
                "Tomadas",
                "Pisos",
                "Executiva",
                "Layout",
                "Demolir e Construir",
                "Apresentação"
            ]
            for option in detalhamento_options:
                if self.detalhamento_vars.get(option, tk.BooleanVar()).get():
                    pdf.cell(0, 5, f"  - {option}", ln=True)

        pdf.ln(5)

        # PROJETOS COMPLEMENTARES
        pdf.set_font("Arial", style="B", size=9)
        pdf.cell(0, 6, "PROJETOS COMPLEMENTARES:", ln=True)
        pdf.set_font("Arial", size=9)
        complementary_options = [
            ("Ar Condicionado", "ar_condicionado"),
            ("Elétrica", "eletrica"),
            ("Dados e Voz", "dados_voz"),
            ("Hidráulica", "hidraulica"),
            ("CFTV", "cftv"),
            ("Alarme", "alarme"),
            ("Incêndio", "incendio")
        ]
        for text, key in complementary_options:
            if self.checks.get(key, tk.BooleanVar()).get():
                pdf.cell(0, 5, f"- {text}", ln=True)

        pdf.ln(5)

        # Salvar o PDF
        pdf.output(pdf_path)
        messagebox.showinfo("Exportar para PDF", f"Dados exportados para {pdf_path} com sucesso.")

    def export_to_excel(self) -> None:
        """Exporta os dados do formulário para um arquivo Excel."""
        if not self.export_path.get():
            messagebox.showerror("Erro", "Selecione uma pasta de exportação antes de continuar.")
            return

        # Nome do arquivo e caminho
        _, excel_path = self._get_export_filename("xlsx")

        # Configuração do Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # INFORMAÇÕES DO CLIENTE
        ws.append(["INFORMAÇÕES DO CLIENTE"])
        client_info = {
            "Nome completo": self.fields.get("nome", ttk.Entry()).get(),
            "Telefone": self.fields.get("telefone", ttk.Entry()).get(),
            "Email": self.fields.get("email", ttk.Entry()).get(),
            "CNPJ/CPF": self.fields.get("cnpj", ttk.Entry()).get(),
            "Responsável legal": self.fields.get("responsavel", ttk.Entry()).get(),
            "Telefone do Responsável": self.fields.get("telefone_responsavel", ttk.Entry()).get(),
            "Endereço": self.fields.get("endereco", ttk.Entry()).get(),
            "CEP": self.fields.get("cep", ttk.Entry()).get()
        }
        for label, value in client_info.items():
            ws.append([label, value])

        # INFORMAÇÕES DO IMÓVEL
        ws.append([])
        ws.append(["INFORMAÇÕES DO IMÓVEL"])
        property_info = {
            "Tipo de Imóvel": self.radio_vars.get("tipo_imovel", tk.StringVar()).get(),
            "Tipo de Construção": self.radio_vars.get("tipo_construcao", tk.StringVar()).get(),
            "Metragem Quadrada": f"{self.fields.get('metragem', ttk.Entry()).get()} m²"
        }
        for label, value in property_info.items():
            ws.append([label, value])

        # PRAZOS DO PROJETO
        ws.append([])
        ws.append(["PRAZOS DO PROJETO"])
        deadlines = {
            "Levantamento": "levantamento",
            "Layout": "layout",
            "Modelagem 3D": "modelagem_3d",
            "Projeto Executivo": "projeto_executivo",
            "Complementares": "complementares"
        }
        for label, key in deadlines.items():
            days_value = self.fields.get(key, ttk.Entry()).get()
            ws.append([label, f"{days_value} DIAS" if days_value else "Não informado"])

        # DEMANDAS DO PROJETO
        ws.append([])
        ws.append(["DEMANDAS DO PROJETO"])
        for nome_entry, descricao_entry in self.demandas_entries:
            nome = nome_entry.get().strip()
            descricao = descricao_entry.get().strip()
            if nome or descricao:
                ws.append([nome, descricao])

        # PROJETO DE ARQUITETURA
        ws.append([])
        ws.append(["PROJETO DE ARQUITETURA"])
        architecture_options = [
            ("Layout", "layout"),
            ("3D", "3d"),
            ("Detalhamento", "detalhamento")
        ]
        for text, key in architecture_options:
            if self.checks.get(key, tk.BooleanVar()).get():
                ws.append([text])

        # Subopções de Detalhamento
        if self.checks.get("detalhamento", tk.BooleanVar()).get():
            detalhamento_options = [
                "Marcenaria",
                "Detalhamento Áreas Molhadas",
                "Forro",
                "Iluminação",
                "Tomadas",
                "Pisos",
                "Executiva",
                "Layout",
                "Demolir e Construir",
                "Apresentação"
            ]
            for option in detalhamento_options:
                if self.detalhamento_vars.get(option, tk.BooleanVar()).get():
                    ws.append([f"  - {option}"])

        # PROJETOS COMPLEMENTARES
        ws.append([])
        ws.append(["PROJETOS COMPLEMENTARES"])
        complementary_options = [
            ("Ar Condicionado", "ar_condicionado"),
            ("Elétrica", "eletrica"),
            ("Dados e Voz", "dados_voz"),
            ("Hidráulica", "hidraulica"),
            ("CFTV", "cftv"),
            ("Alarme", "alarme"),
            ("Incêndio", "incendio")
        ]
        for text, key in complementary_options:
            if self.checks.get(key, tk.BooleanVar()).get():
                ws.append([text])

        # Salvar o Excel
        wb.save(excel_path)
        messagebox.showinfo("Exportar para Excel", f"Dados exportados para {excel_path} com sucesso.")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Formulário de Projeto Arquitetônico")
    root.geometry("1000x700")
    root.minsize(800, 600)

    app_logic = AppLogic()
    app_logic.set_root(root)
    ui = UI(root, app_logic)

    root.mainloop()
